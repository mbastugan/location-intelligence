"""Load SERPAVI municipal median rent (€/m²) for MVP cities into SQLite."""

from __future__ import annotations

import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from pipeline import DB_PATH, DATA_DIR, connect
from pipeline.jobs.init_db import derive_gross_yield, main as init_db_main
from pipeline.scoring import compute_and_store_scores

SERPAVI_PATH = DATA_DIR / "raw" / "serpavi" / "serpavi_2011_2024.xlsx"
SERPAVI_URL = (
    "https://cdn.mivau.gob.es/portal-web-mivau/vivienda/serpavi/"
    "2026-03_09_bd_SERPAVI_2011-2024%20-%20DEFINITIVO%20WEB_v2.xlsx"
)

# INE municipality codes used in location.admin_code
TARGET_CUMUN = {
    "29067": "malaga",
    "46250": "valencia",
    "03014": "alicante",
}

# Collective housing (vivienda colectiva) median €/m²/month
RENT_M2_COL_PREFIX = "ALQM2_LV_M_VC_"


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def ensure_serpavi_source(conn) -> int:
    row = conn.execute(
        "SELECT id FROM data_source WHERE code = 'serpavi_mivau'"
    ).fetchone()
    if row:
        return int(row["id"])
    conn.execute(
        """
        INSERT INTO data_source (code, name, homepage_url, license_note, notes)
        VALUES (
          'serpavi_mivau',
          'SERPAVI (MIVAU)',
          'https://www.mivau.gob.es/vivienda/alquila-bien-es-tu-derecho/serpavi',
          'Official Spanish open statistical reuse — attribute MIVAU/SERPAVI',
          'Tax-based habitual rent indicators; annual lag'
        )
        """
    )
    conn.commit()
    return int(
        conn.execute(
            "SELECT id FROM data_source WHERE code = 'serpavi_mivau'"
        ).fetchone()["id"]
    )


def parse_year_columns(headers: list) -> dict[int, int]:
    """Map year -> column index for median collective rent €/m²."""
    mapping: dict[int, int] = {}
    for idx, name in enumerate(headers):
        if not isinstance(name, str):
            continue
        match = re.fullmatch(rf"{RENT_M2_COL_PREFIX}(\d{{2}})", name)
        if not match:
            continue
        year = 2000 + int(match.group(1))
        mapping[year] = idx
    return mapping


def load_municipal_rents(path: Path) -> dict[str, dict[int, float]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Municipios"]
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))
    year_cols = parse_year_columns(headers)
    try:
        cumun_idx = headers.index("CUMUN")
    except ValueError as exc:
        raise RuntimeError("CUMUN column missing in SERPAVI Municipios sheet") from exc

    out: dict[str, dict[int, float]] = {code: {} for code in TARGET_CUMUN}
    for row in rows:
        if not row:
            continue
        cumun = row[cumun_idx]
        if cumun is None:
            continue
        code = str(cumun).zfill(5)
        if code not in TARGET_CUMUN:
            continue
        for year, col in year_cols.items():
            value = row[col]
            if value is None:
                continue
            try:
                out[code][year] = float(value)
            except (TypeError, ValueError):
                continue
    wb.close()
    return out


def upsert_rents(conn, rents: dict[str, dict[int, float]], source_id: int) -> int:
    metric_id = conn.execute(
        "SELECT id FROM metric_definition WHERE code = 'rent_m2'"
    ).fetchone()["id"]
    seed_source = conn.execute(
        "SELECT id FROM data_source WHERE code = 'seed_manual'"
    ).fetchone()
    locations = {
        row["admin_code"]: int(row["id"])
        for row in conn.execute(
            "SELECT id, admin_code FROM location WHERE admin_code IS NOT NULL"
        ).fetchall()
    }
    collected = utc_now_iso()
    count = 0
    for cumun, by_year in rents.items():
        loc_id = locations.get(cumun)
        if loc_id is None:
            continue
        if seed_source:
            conn.execute(
                """
                DELETE FROM metric_observation
                WHERE location_id = ? AND metric_id = ? AND source_id = ?
                """,
                (loc_id, metric_id, int(seed_source["id"])),
            )
        for year, value in sorted(by_year.items()):
            period = f"{year}-01-01"
            period_end = f"{year}-12-31"
            conn.execute(
                """
                INSERT INTO metric_observation
                  (location_id, metric_id, source_id, period_start, period_end, value,
                   currency_code, quality_flag, source_url, collected_at)
                VALUES (?, ?, ?, ?, ?, ?, 'EUR', 'ok', ?, ?)
                ON CONFLICT(location_id, metric_id, period_start, source_id) DO UPDATE SET
                  period_end=excluded.period_end,
                  value=excluded.value,
                  quality_flag=excluded.quality_flag,
                  source_url=excluded.source_url,
                  collected_at=excluded.collected_at
                """,
                (
                    loc_id,
                    metric_id,
                    source_id,
                    period,
                    period_end,
                    round(value, 3),
                    SERPAVI_URL,
                    collected,
                ),
            )
            count += 1
    conn.commit()
    return count


def main() -> None:
    if not SERPAVI_PATH.exists():
        raise FileNotFoundError(
            f"Missing {SERPAVI_PATH}. Download from {SERPAVI_URL}"
        )
    if not DB_PATH.exists():
        init_db_main()

    print(f"Reading {SERPAVI_PATH} ...")
    rents = load_municipal_rents(SERPAVI_PATH)
    for code, slug in TARGET_CUMUN.items():
        years = sorted(rents.get(code, {}))
        latest = rents.get(code, {}).get(years[-1]) if years else None
        print(f"  {slug} ({code}): {len(years)} years, latest={latest}")

    conn = connect()
    source_id = ensure_serpavi_source(conn)
    n = upsert_rents(conn, rents, source_id)
    # Prefer SERPAVI for latest rent when deriving yield: remove provisional seed rents
    # for overlapping periods so latest-value queries prefer official source years.
    # Keep seed property prices; recompute yield from latest matching year pairs.
    derive_gross_yield(conn)
    compute_and_store_scores(conn)
    conn.close()
    print(f"Upserted {n} SERPAVI rent observations")


if __name__ == "__main__":
    main()
